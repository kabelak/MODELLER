__author__ = 'Kavin'

import sys
import xlrd
from openpyxl import *


def main(xlfile):
    wb = xlrd.open_workbook(xlfile)
    ws = wb.sheet_by_name('Sheet1')

    workbook = Workbook()
    sheet = workbook.active

    curr_row = 0
    ww = 1
    while curr_row < ws.nrows:
        # print(ws.cell_value(curr_row, 2))
        if (ws.cell_value(curr_row, 2) != "" and ws.cell_value(curr_row, 2) != "AVRG"):

            if (ws.cell_value(curr_row, 2) % 2 != 0):
                for i in range(2, 7, 1):
                    sheet.cell(row=ww, column=(i + 1)).value = ws.cell_value(curr_row, i)

            if (ws.cell_value(curr_row, 2) % 2 == 0):
                for i in range(9, 14, 1):
                    sheet.cell(row=ww, column=(i + 1)).value = ws.cell_value(curr_row, (i - 7))
                ww += 1

                #print(str(ws.cell_value(curr_row, 2)) + " " + str(curr_row))
        else:
            for i in range(1, ws.ncols, 1):
                sheet.cell(row=ww, column=i).value = ws.cell_value(curr_row, i)
            ww += 1

        curr_row += 1


    # sheet['F5'] = 3.14
    #sheet.cell(row=1, column=2).value = 1.22


    workbook.save('output.xlsx')


if __name__ == "__main__":
    main(sys.argv[1])
