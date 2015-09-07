__author__ = 'Kavin'

import sys
import re
from openpyxl import *


def main(xlfile):
    # Create objects for source and destination worksheets
    wb_from = load_workbook(xlfile, data_only=True)
    wb_to = Workbook()

    # Iterate over worksheets in source, creating mirror copy names in destination
    for ws_from in wb_from.worksheets:
        ws_to = wb_to.create_sheet(title=ws_from._title)

        # Reset row numbering for read(curr_row) and write (ww)
        curr_row = 1
        ww = 1
        # Iterate until last row in source
        while curr_row < ws_from.get_highest_row():
            # Look for data
            if ws_from.cell(row=curr_row, column=3).value is not None and ws_from.cell(row=curr_row,
                                                                                       column=3).value != "AVRG":
                # Copy odd-numbered rows directly
                if (ws_from.cell(row=curr_row, column=3).value % 2 != 0):
                    for i in range(3, 9, 1):
                        ws_to.cell(row=ww, column=i).value = ws_from.cell(row=curr_row, column=i).value

                # Transpose even-numbered rows next to above odd-numbered row
                if (ws_from.cell(row=curr_row, column=3).value % 2 == 0):
                    for i in range(10, 16, 1):
                        ws_to.cell(row=ww, column=i).value = ws_from.cell(row=curr_row, column=(i - 7)).value

                    # After an odd-even cycle, increase write-row counter
                    ww += 1

            # For title and other rows, copy as is
            else:
                # Only difference is for row with AVRG, where we do not want to copy the adjacent value
                if ws_from.cell(row=curr_row, column=3).value == "AVRG":
                    for i in range(1, 4, 1):
                        ws_to.cell(row=ww, column=i).value = ws_from.cell(row=curr_row, column=i).value
                        ws_to.cell(row=ww, column=i).font = ws_from.cell(row=curr_row, column=i).font

                else:
                    for i in range(1, 16, 1):
                        ws_to.cell(row=ww, column=i).value = ws_from.cell(row=curr_row, column=i).value
                        ws_to.cell(row=ww, column=i).font = ws_from.cell(row=curr_row, column=i).font
                # Increase write-row counter after each non-data row
                ww += 1

            curr_row += 1


        # Create output file
        fname = re.search('(.*)\.(\w*)', xlfile)
        fname2 = str(fname.group(1)) + '_reformatted.' + str(fname.group(2))
        # out = open(fname2, "w")

        # Save new workbook
        wb_to.save(fname2)

        # Delete empty first "Sheet"
        wb_del = load_workbook(fname2)
        ws_del = wb_del.get_sheet_by_name("Sheet")
        wb_del.remove_sheet(ws_del)

        wb_del.save(fname2)


if __name__ == "__main__":
    main(sys.argv[1])
