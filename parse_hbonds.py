__author__ = 'Kavin'
__usage__ = ''

import sys
from openpyxl import *
from pathlib import Path
import re
import os


def parseline(_line):
    hb_cols = re.search('^(A\S*)\s+(\S*)\s+(\S*)\s+(\S*)\s+(\S*)\s+(\S*)\s+.*$', _line)
    if hb_cols:
        _ara = str(hb_cols.group(1)[0:3] + hb_cols.group(1)[-3:])
        _donor_re = re.search('(\w{3})_(\d{0,3})(@.*)', hb_cols.group(3))
        _res = _donor_re.group(1)
        _resnum = int(_donor_re.group(2)) + 43
        _resatom = _donor_re.group(3)
        _fraction = float(hb_cols.group(5))

        #   Cleanup the residue names
        if _res == 'HID':
            _res = 'HIS'
        elif _res == 'HIE':
            _res = 'HIS'

        return _ara, _res, _resnum, _resatom, _fraction


def main(hbondsheet):
    #   A bit of naming maintenace, to be used throughout
    fname = re.search('.*/(.*)(_.*)\.(\w*)', hbondsheet)
    fname2 = str(fname.group(1)) + '_analysed.xlsx'

    #   Create a workbook and set row 1
    if os.path.exists(fname2):
        wb_to = load_workbook(fname2)
        ws_to = wb_to.create_sheet(title=(str(fname.group(1)) + str(fname.group(2)))[-30:])
    else:
        wb_to = Workbook()
        ws_to = wb_to.create_sheet(title=(str(fname.group(1)) + str(fname.group(2)))[-30:])
    _row = 1

    #   Create data structure to count to add the fractions
    _counts = {}

    #   Open the hbonds file and iterate by line
    with open(hbondsheet, "r") as hb:
        for _line in hb.readlines():
            try:
                #   Pick up the relevant data from each line, add to row
                _ara, _res, _resnum, _resatom, _fraction = parseline(_line)
                ws_to.cell(row=_row, column=1).value = _ara
                ws_to.cell(row=_row, column=2).value = _res + ' ' + str(_resnum) + _resatom
                ws_to.cell(row=_row, column=3).value = _fraction

                #   Add up fraction for each residue
                _donor = str(_res + ' ' + str(_resnum))
                if not _donor in _counts:
                    _counts[_donor] = _fraction
                else:
                    _counts[_donor] += _fraction

                    # Iterate the row counter and go to next line
                _row += 1

            except TypeError:
                pass

    # Create entries for the counted stats
    _row = 1
    for key in sorted(_counts, key=_counts.get, reverse=True):
        if round(_counts[key], 4) > 0.0001:
            ws_to.cell(row=_row, column=5).value = key
            ws_to.cell(row=_row, column=6).value = round(_counts[key], 4)
        _row += 1
        # print(key, 'corresponds to', round(_counts[key],4))

    #   Create output file
    wb_to.save(fname2)

    wb_to.close()
    #   Delete empty first "Sheet"
    wb_del = load_workbook(fname2)
    if 'Sheet' in wb_del.sheetnames:
        ws_del = wb_del.get_sheet_by_name("Sheet")
        wb_del.remove_sheet(ws_del)
        wb_del.save(fname2)

    wb_del.close()
    print("created " + fname2)

if __name__ == "__main__":
    main(sys.argv[1])
