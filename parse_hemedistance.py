__author__ = 'Kavin'
__usage__ = ''

import sys
import re
import argparse
import os
from openpyxl import *

def check_range(arg):  # Function to ensure correct range of IRE Length is input at the command prompt
    try:
        value = int(arg)
    except ValueError as err:
        raise argparse.ArgumentTypeError(str(err))
    if value < 1 or value > 30:
        message = "Expected length between 1 and 30 inclusive, received value = {}".format(value)
        raise argparse.ArgumentTypeError(message)
    return value


def namerows(_sheet):
    #   Create row headings
    _sheet.cell(row=1, column=1).value = "Pose"
    _sheet.cell(row=2, column=1).value = "Repeat"
    _sheet.cell(row=3, column=1).value = "Omega 6"
    _sheet.cell(row=4, column=1).value = "Omega 9"
    _sheet.cell(row=5, column=1).value = "Omega 12"
    _sheet.cell(row=6, column=1).value = "Omega 15"
    _sheet.cell(row=7, column=1).value = "Carbon 19"
    _sheet.cell(row=8, column=1).value = "Carboxylic head"

    return _sheet


def main(argv):
    #   Parse input arguments
    parser = argparse.ArgumentParser(description='Parses Fe-Omega/Carbon Atom distance files',
                                     usage='python3 parse_hemedistance.py <file> <cutoff>')
    parser.add_argument('file', help='Distance file')
    parser.add_argument('cutoff', type=check_range, help='Cut-off value',
                        default=6, nargs=1)
    args = parser.parse_args()

    #   Check that file is input
    if not args.file:
        parser.error('Please specify a distance file')

    with open(args.file, "r") as df:

        #   Introduce counters
        total_frames = 0
        relevant_frames = 0

        #   Process file line by line
        for _line in df.readlines():
            dist_cols = re.search('^\s*(\d+)\.000\s*(\d+\.\d+)', _line)
            if dist_cols:
                total_frames += 1
                if float(dist_cols.group(2)) < args.cutoff[0]:
                    relevant_frames += 1

                    #      print('\tFile Name:\t', args.file, '\n',
                    #           '\tTotal Frames:\t', total_frames, '\n',
                    #          '\tCutoff Value:\t', args.cutoff[0], '\n',
                    #         '\t< cutoff :\t', relevant_frames, '\n',
                    #        '\t% < cutoff:\t', round(relevant_frames / total_frames * 100, 2))

    # A bit of naming maintenance, to be used throughout
    fname = re.search('(Mut_R111A_R117A).*/Distance_FEto(.*)\.(\d)\.(.*)\.agr', os.path.abspath(args.file))
    fname2 = str(fname.group(1)) + '_hemedistance_analysed.xlsx'

    print(fname.group(1), fname.group(2), fname.group(3), fname.group(4))
    # main_name = fname.group(1)
    bond = fname.group(2)
    pose = fname.group(3)
    repeat = fname.group(4)

    # print(os.path.abspath(args.file))

    #   Create a workbook and set row 1
    if os.path.exists(fname2):
        wb = load_workbook(fname2)
        if not str(args.cutoff[0]) in wb.sheetnames:
            ws = wb.create_sheet(title=(str(args.cutoff[0])))
            ws = namerows(ws)
            _col = 2
        else:
            ws = wb[str(args.cutoff[0])]
            # _row = ws.max_rows
            _col = ws.max_column

    else:
        wb = Workbook()
        ws = wb.create_sheet(title=(str(args.cutoff[0])))
        ws = namerows(ws)
        _col = 2

    if not ws.cell(row=1, column=_col).value:
        ws.cell(row=1, column=_col).value = pose

    if not ws.cell(row=2, column=_col).value:
        ws.cell(row=2, column=_col).value = str(repeat)
    elif ws.cell(row=2, column=_col).value != str(repeat) and bond == "Omega6":
        _col += 1

    if bond == "Omega6":
        ws.cell(row=3, column=_col).value = round(relevant_frames / total_frames * 100, 2)
    elif bond == "Omega9":
        ws.cell(row=4, column=_col).value = round(relevant_frames / total_frames * 100, 2)
    elif bond == "Omega12":
        ws.cell(row=5, column=_col).value = round(relevant_frames / total_frames * 100, 2)
    elif bond == "Omega15":
        ws.cell(row=6, column=_col).value = round(relevant_frames / total_frames * 100, 2)
    elif bond == "C19":
        ws.cell(row=7, column=_col).value = round(relevant_frames / total_frames * 100, 2)
    elif bond == "CO2":
        ws.cell(row=8, column=_col).value = round(relevant_frames / total_frames * 100, 2)
    # Save and exit, clearing the 'Sheet' worksheet
    wb.save(fname2)
    wb.close()
    #   Delete empty first "Sheet"
    wb_del = load_workbook(fname2)
    if 'Sheet' in wb_del.sheetnames:
        ws_del = wb_del.get_sheet_by_name("Sheet")
        wb_del.remove_sheet(ws_del)
        wb_del.save(fname2)

    wb_del.close()


if __name__ == "__main__":
    main(sys.argv)

'''
    #   Create row headings
    if ws.cell(row=_row, column=_col).value:
        _row += 1
        ws.cell(row=_row, column=_col).value = bond
    else:
        ws.cell(row=_row, column=_col).value = bond

'''
