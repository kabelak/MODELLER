__author__ = 'Kavin'
__usage__ = ''

import sys
import re
import argparse
import os
from openpyxl import *
from rpy2.robjects import r


def namerows(_sheet):
    #   Create row headings
    _sheet.cell(row=1, column=1).value = "Pose"
    _sheet.cell(row=2, column=1).value = "Repeat"
    _sheet.cell(row=3, column=1).value = "Stats"
    _sheet.cell(row=4, column=1).value = "Omega 6"
    _sheet.cell(row=5, column=1).value = "Omega 9"
    _sheet.cell(row=6, column=1).value = "Omega 12"
    _sheet.cell(row=7, column=1).value = "Omega 15"
    _sheet.cell(row=8, column=1).value = "Carbon 19"
    _sheet.cell(row=9, column=1).value = "Carboxylic head"

    return _sheet


def main(argv):
    #   Parse input arguments
    parser = argparse.ArgumentParser(description='Parses Fe-Omega/Carbon Atom distance files',
                                     usage='python3 parse_hemedistance.py <file> <cutoff>')
    parser.add_argument('file', help='Distance file')
    args = parser.parse_args()

    #   Check that file is input
    if not args.file:
        parser.error('Please specify a distance file')

    # A bit of naming maintenance, to be used throughout
    fname = re.search('(Mut_R117A).*/Distance_FEto(.*)\.(\d)\.(.*)\.agr', os.path.abspath(args.file))
    fname2 = str(fname.group(1)) + '_hemedistance_stats.xlsx'

    # Some output to stdin to show progress when looping
    print(fname.group(1), fname.group(2), fname.group(3), fname.group(4))
    bond = fname.group(2)
    pose = fname.group(3)
    repeat = fname.group(4)

    #   Create a workbook and set row 1
    if os.path.exists(fname2):
        wb = load_workbook(fname2)
        if not 'Stats' in wb.sheetnames:
            ws = wb.create_sheet(title='Stats')
            ws = namerows(ws)
            _col = 2
        else:
            ws = wb['Stats']
            # _row = ws.max_rows
            if bond == "Omega6":
                _col = ws.max_column
            else:
                _col = ws.max_column - 3

    else:
        wb = Workbook()
        ws = wb.create_sheet(title='Stats')
        ws = namerows(ws)
        _col = 2

    # Pass file to R and process as table
    r.assign('dataf', os.path.abspath(args.file))
    r('dc <- read.table(dataf, skip=8, col.names = c("Time", "Distance"))')
    r('d <- dc[,"Distance"]')

    # Logics to ensure new file is correctly written, or existing file
    # correctly updated
    if not ws.cell(row=1, column=_col).value:
        ws.cell(row=1, column=_col).value = pose

    if not ws.cell(row=2, column=_col).value:
        ws.cell(row=2, column=_col).value = str(repeat)
    elif ws.cell(row=2, column=_col).value != str(repeat) and bond == "Omega6":
        _col += 1

    # Parse data and populate cells with stats
    if bond == "Omega6":
        ws.cell(row=2, column=_col + 1).value = str(repeat)
        ws.cell(row=2, column=_col + 2).value = str(repeat)
        ws.cell(row=2, column=_col + 3).value = str(repeat)
        ws.cell(row=1, column=_col + 1).value = pose
        ws.cell(row=1, column=_col + 2).value = pose
        ws.cell(row=1, column=_col + 3).value = pose
        ws.cell(row=3, column=_col).value = 'Min'
        ws.cell(row=3, column=_col + 1).value = 'Max'
        ws.cell(row=3, column=_col + 2).value = 'Mean'
        ws.cell(row=3, column=_col + 3).value = 'Median'
        ws.cell(row=4, column=_col).value = r('min(d)')[0]
        ws.cell(row=4, column=_col + 1).value = r('max(d)')[0]
        ws.cell(row=4, column=_col + 2).value = r('mean(d)')[0]
        ws.cell(row=4, column=_col + 3).value = r('median(d)')[0]
    elif bond == "Omega9":
        ws.cell(row=5, column=_col).value = r('min(d)')[0]
        ws.cell(row=5, column=_col + 1).value = r('max(d)')[0]
        ws.cell(row=5, column=_col + 2).value = r('mean(d)')[0]
        ws.cell(row=5, column=_col + 3).value = r('median(d)')[0]
    elif bond == "Omega12":
        ws.cell(row=6, column=_col).value = r('min(d)')[0]
        ws.cell(row=6, column=_col + 1).value = r('max(d)')[0]
        ws.cell(row=6, column=_col + 2).value = r('mean(d)')[0]
        ws.cell(row=6, column=_col + 3).value = r('median(d)')[0]
    elif bond == "Omega15":
        ws.cell(row=7, column=_col).value = r('min(d)')[0]
        ws.cell(row=7, column=_col + 1).value = r('max(d)')[0]
        ws.cell(row=7, column=_col + 2).value = r('mean(d)')[0]
        ws.cell(row=7, column=_col + 3).value = r('median(d)')[0]
    elif bond == "C19":
        ws.cell(row=8, column=_col).value = r('min(d)')[0]
        ws.cell(row=8, column=_col + 1).value = r('max(d)')[0]
        ws.cell(row=8, column=_col + 2).value = r('mean(d)')[0]
        ws.cell(row=8, column=_col + 3).value = r('median(d)')[0]
    elif bond == "CO2":
        ws.cell(row=9, column=_col).value = r('min(d)')[0]
        ws.cell(row=9, column=_col + 1).value = r('max(d)')[0]
        ws.cell(row=9, column=_col + 2).value = r('mean(d)')[0]
        ws.cell(row=9, column=_col + 3).value = r('median(d)')[0]

    # Save and exit, clearing the 'Sheet' worksheet
    wb.save(fname2)
    wb.close()
    #   Delete empty first "Sheet"
    wb_del = load_workbook(fname2)
    if 'Sheet' in wb_del.sheetnames:
        ws_del = wb_del.get_sheet_by_name("Sheet")
        wb_del.remove_sheet(ws_del)
        wb_del.save(fname2)

    wb_del.close()


if __name__ == "__main__":
    main(sys.argv)

